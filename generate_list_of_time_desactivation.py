import os
import openpyxl

from openpyxl import load_workbook
from openpyxl.styles import Alignment
ERI_dose_control = load_workbook('C:\\Work\\Python\\test2\\Для графика высвечивания.xlsx')
sheets = ERI_dose_control.sheetnames
worksheet_main = ERI_dose_control[sheets[1]]
worksheet_result = ERI_dose_control[sheets[2]]

from openpyxl import load_workbook
ERI_date_with_fluence = load_workbook('C:\\Work\\Python\\test2\\Результат.xlsx')
worksheet_list_of_ERI = ERI_date_with_fluence.active

#Составляем для сет для анализа
tuple_ws = worksheet_list_of_ERI['A1':'E137']
list_of_ERI = []
for i in range (0, len(tuple_ws)):
    if tuple_ws[i][0].value not in list_of_ERI:
        list_of_ERI.append(tuple_ws[i][0].value)

#Ищем ЭРИ в списке высвечивания
current_row = 2
current_column = 1
for ERI in list_of_ERI:
    for row in range(4, 217):
        if worksheet_main.cell(row=row, column=1).value == ERI:
            worksheet_result.cell(row=current_row, column=1, value=ERI)
            worksheet_result.cell(row=current_row, column=2,
                                  value=worksheet_main.cell(row=row, column=6).value)
            current_column = 7
            for i in range(7, 25):
                if worksheet_main.cell(row=row, column=i).value != None:
                    date = worksheet_main.cell(row=1, column=i).value.strftime('%d.%m.%y')
                    radioactivity = worksheet_main.cell(row=row, column=i).value
                    worksheet_result.cell(row=current_row, column=current_column,
                                          value=f'{date}\\\n{radioactivity}')
                    worksheet_result.cell(row=current_row, column=current_column).alignment = Alignment(wrapText=True)
                    current_column += 1
            current_row += 1
    for row in range(1, 137):
        if worksheet_list_of_ERI.cell(row=row, column=1).value == ERI:
            req_fluence = worksheet_list_of_ERI.cell(row=row, column=3).value
            act_fluence = worksheet_list_of_ERI.cell(row=row, column=4).value
            name_file = worksheet_list_of_ERI.cell(row=row, column=5).value
            date_pulse = worksheet_list_of_ERI.cell(row=row, column=2).value
            worksheet_result.cell(row=current_row, column=3, value=req_fluence)
            worksheet_result.cell(row=current_row, column=4, value=act_fluence)
            worksheet_result.cell(row=current_row, column=5, value=name_file)
            worksheet_result.cell(row=current_row, column=6, value=f'{date_pulse}\\\n0')
            worksheet_result.cell(row=current_row, column=6).alignment = Alignment(wrapText=True)
            current_row += 1
    current_row += 1
ERI_dose_control.save('C:\\Work\\Python\\test2\\Для графика высвечивания.xlsx')
ERI_dose_control.close()



