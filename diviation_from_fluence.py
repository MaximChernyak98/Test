import openpyxl

from openpyxl import load_workbook
from openpyxl.styles import Alignment

wb = load_workbook('C:\\Work\\Python\\test2\\Для графика высвечивания.xlsx')
ws_data = wb["Сведенные с датой импульса"]
ws_result = wb["Отличие во флюенсе"]

for i in range(2, 131):
    req_fluence = ws_data.cell(row=i, column=3).value
    act_fluence = ws_data.cell(row=i, column=4).value
    diviation = (act_fluence - req_fluence)/req_fluence * 100
    ws_result.cell(row=i, column=1, value=req_fluence)
    ws_result.cell(row=i, column=1).number_format = '0.00E+00'
    ws_result.cell(row=i, column=2, value=act_fluence)
    ws_result.cell(row=i, column=2).number_format = '0.00E+00'
    ws_result.cell(row=i, column=3, value=diviation)
    ws_result.cell(row=i, column=3).number_format = '0'

wb.save('C:\\Work\\Python\\test2\\Для графика высвечивания.xlsx')
wb.close()