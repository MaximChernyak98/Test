import openpyxl

from openpyxl import load_workbook
from openpyxl.styles import Alignment
from datetime import datetime, date
from decimal import Decimal


wb = load_workbook('C:\\Work\\Python\\test2\\Для графика высвечивания.xlsx')
ws_data = wb["Сведенные с датой импульса"]
ws_result = wb["Время высвечивания"]
ws_graphs = wb["Графики высвечивания"]
current_column = 1

def find_last_date (sheet, row):
    last_date = ''
    for column in range (7, 20):
        date = sheet.cell(row=row, column=column).value
        if date != None:
            last_date = date.split('\\')[0]
    return last_date

def return_date_and_radioactivity (cell):
    text = cell.value
    list = text.split('\\')
    radioactivity = float(list[1].strip('\n'))
    date_list = list[0].split('.')
    date_pulse = date(int(date_list[2]), int(date_list[1]), int(date_list[0]))
    return date_pulse, radioactivity

for row in range(2, 131):
    ERI = ws_data.cell(row=row, column=1).value
    fluence = ws_data.cell(row=row, column=4).value
    num_of_measure = 0
    for i in range(7, 20):
        if ws_data.cell(row=row, column=i).value != None:
            num_of_measure += 1

    date_pulse, radioactivity = return_date_and_radioactivity(ws_data.cell(row=row, column=6))
    ws_result.cell(row=row, column=1, value=ERI)
    ws_result.cell(row=row, column=2, value=fluence)
    ws_result.cell(row=row, column=2).number_format = '0.00E+00'
    date_of_control, radioactivity = return_date_and_radioactivity(ws_data.cell(row=row, column=6 + num_of_measure))
    if radioactivity < 0.2:
        ws_result.cell(row=row, column=3, value=(date_of_control - date_pulse).days)
        ws_result.cell(row=row, column=3).number_format = '0'
        ws_result.cell(row=row, column=4, value='Высветился')
    else:
        ws_result.cell(row=row, column=3, value=(date_of_control - date_pulse).days)
        ws_result.cell(row=row, column=3).number_format = '0'
        ws_result.cell(row=row, column=4, value=f'Еще светится, остаточная {radioactivity} мкЗв/ч')

    if num_of_measure > 1:
        ws_graphs.cell(row=1, column=(current_column + 1), value=(ERI + ', ' + ('%.2E' % Decimal(str(fluence)))))
        for i in range (7, (7 + num_of_measure)):
            date_of_control, radioactivity = return_date_and_radioactivity(ws_data.cell(row=row, column=i))
            ws_graphs.cell(row=(2 - 7 + i), column=current_column, value=(date_of_control - date_pulse).days)
            ws_graphs.cell(row=(2 - 7 + i), column=current_column).number_format = '0'
            ws_graphs.cell(row=(2 - 7 + i), column=(current_column + 1), value=radioactivity)
        current_column += 3

wb.save('C:\\Work\\Python\\test2\\Для графика высвечивания.xlsx')
wb.close()



